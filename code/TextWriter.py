#! python3

import sys

import logging
logging.basicConfig(level = logging.DEBUG, format = '[%(levelname)s][%(asctime)s]{ %(message)s }')

vArgArray: list[str] = sys.argv
vArgCount: int = len(vArgArray) - 1
logging.debug('the argument array is:\n' + str(vArgArray))
if vArgCount != 2:
    raise Exception('2 command line arguments required: input and output paths; we have ' + str(vArgCount))

import json
vSourcePath: str = vArgArray[1]
vSourceFile = open(vSourcePath, 'r', encoding = 'utf-8')
vSourceJson = json.load(vSourceFile)
vSourceFile.close()

import csv
import openpyxl
vOutputPath: str = vArgArray[2]
vOutputFile = open(vOutputPath, 'w', encoding = 'utf-8')
vOutputCsv = csv.writer(vOutputFile, dialect = 'excel')
vOutputCsv.writerow(('место', 'текст'))
vOutputWb = openpyxl.Workbook()
vOutputWs = vOutputWb.active
if not vOutputWs:
    raise Exception('failed to get the sheet')
vOutputWs.cell(row = 1, column = 1).value = 'место'
vOutputWs.cell(row = 1, column = 2).value = 'текст'
vOutputIndex: int = 2
for vSourceItem, vSourceText in vSourceJson.items():
    vOutputCsv.writerow((vSourceItem, vSourceText))
    vOutputWs.cell(row = vOutputIndex, column = 1).value = vSourceItem
    vOutputWs.cell(row = vOutputIndex, column = 2).value = vSourceText
    vOutputIndex += 1
vOutputFile.close()
vOutputWb.save(vOutputPath + '.xlsx')
